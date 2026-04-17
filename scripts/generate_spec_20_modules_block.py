"""Спецификация 20 модулей в формате «Блок спецификации»: две строки на профиль (два подреза).

Код профиля: «СК-{d}-{L}», «СС-{d}-{L}», «Р-{d}-{L}» — d = шифр серии (0…3 → Н20…Н23),
L = длина в мм (совпадает с колонкой «Длина»).
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

COL0 = 6
HEADER_ROW = 13
HEADERS = [
    "№ п/п",
    "Наименование",
    "Тип профиля",
    "Длина",
    "Угол запила",
    "Сторона запила",
    "Количество",
    "QR-код",
]

# (каркас NordFox?, префикс или имя, цифра 0…3 если каркас, базовая длина мм, угол1, угол2)
PIECE_BLUEPRINTS: list[tuple[bool, str, int, int, int, int]] = [
    (True, "СК", 1, 1045, 45, 87),
    (True, "СК", 0, 1050, 90, 110),
    (True, "Р", 1, 3030, 70, 65),
    (True, "Р", 2, 3028, 90, 77),
    (True, "СС", 1, 456, 90, 88),
    (True, "СС", 1, 457, 90, 87),
    (True, "СС", 2, 458, 90, 85),
    (True, "СС", 3, 459, 90, 55),
    (True, "СС", 0, 451, 90, 67),
    (False, "L15", 0, 500, 90, 80),
    (False, "DT21", 0, 1000, 88, 77),
    (True, "СК", 3, 1280, 62, 58),
]


def _length_mm(base: int, mod_i: int, j: int) -> int:
    return base + (mod_i * 11 + j * 7) % 97


def _profile_code(is_nordfox: bool, prefix_or_name: str, digit: int, length_mm: int) -> str:
    if not is_nordfox:
        return prefix_or_name
    return f"{prefix_or_name}-{digit}-{length_mm}"


def main() -> None:
    out = ROOT / "test" / "spec_20_modules_block.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None

    for j, h in enumerate(HEADERS):
        ws.cell(HEADER_ROW, COL0 + j, h)

    r = HEADER_ROW + 1
    item_no = 0
    for mod_i in range(1, 21):
        name = f"Модуль М{mod_i}"
        for j, row in enumerate(PIECE_BLUEPRINTS):
            item_no += 1
            is_nf, pfx, digit, base_len, a1, a2 = row
            L = _length_mm(base_len, mod_i, j)
            prof = _profile_code(is_nf, pfx, digit, L)
            ws.cell(r, COL0, item_no)
            ws.cell(r, COL0 + 1, name)
            ws.cell(r, COL0 + 2, prof)
            ws.cell(r, COL0 + 3, L)
            ws.cell(r, COL0 + 4, a1)
            ws.cell(r, COL0 + 5, "Правая")
            ws.cell(r, COL0 + 6, 1)
            r += 1
            ws.cell(r, COL0 + 4, a2)
            ws.cell(r, COL0 + 5, "Левая")
            r += 1

    wb.save(out)
    print(out)


if __name__ == "__main__":
    main()
