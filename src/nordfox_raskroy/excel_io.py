from __future__ import annotations

import re
import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook

from nordfox_raskroy.models import SpecRow
logger = logging.getLogger("nordfox_raskroy.excel_io")

HEADER_ALIASES = {
    "№ п/п": "item_no",
    "Наименование": "module_name",
    "Тип профиля": "profile_code",
    "Длина": "length_mm",
    "Угол запила": "cut_angle",
    "Количество": "quantity",
    "QR-код": "qr",
}

# Доп. заголовки шаблона «Блок спецификации» (не используются в write_spec_workbook)
HEADER_ALIASES_EXTRA = {
    "Сторона запила": "cut_side",
}


@dataclass(frozen=True)
class ParseStats:
    parsed_rows: int
    skipped_blank_rows: int
    skipped_invalid_rows: int


def _find_header_row(ws, max_scan: int = 60) -> int:
    for r in range(1, max_scan + 1):
        for c in range(1, 15):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() == "Наименование":
                return r
    raise ValueError("Не найдена строка заголовка с колонкой «Наименование»")


def _map_headers(ws, header_row: int) -> dict[str, int]:
    col_by_key: dict[str, int] = {}
    for c in range(1, 32):
        raw = ws.cell(header_row, c).value
        if raw is None:
            continue
        label = str(raw).strip()
        if label in HEADER_ALIASES:
            col_by_key[HEADER_ALIASES[label]] = c
        elif label in HEADER_ALIASES_EXTRA:
            col_by_key[HEADER_ALIASES_EXTRA[label]] = c
    required = {"module_name", "profile_code", "length_mm", "cut_angle", "quantity"}
    missing = required - set(col_by_key)
    if missing:
        raise ValueError(f"Не хватает колонок: {missing}")
    return col_by_key


def _cell_str(ws, r: int, c: int | None) -> str | None:
    if c is None:
        return None
    v = ws.cell(r, c).value
    if v is None:
        return None
    return str(v).strip()


def _cell_int(ws, r: int, c: int) -> int:
    v = ws.cell(r, c).value
    if v is None:
        raise ValueError(f"Пустая ячейка (строка {r}, колонка {c})")
    if isinstance(v, bool):
        raise ValueError("Некорректное число")
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        return int(round(v))
    s = str(v).strip().replace(" ", "").replace(",", ".")
    if not s:
        raise ValueError("Пустое число")
    return int(round(float(s)))


def _is_second_angle_row(ws, r: int, cols: dict[str, int]) -> bool:
    """
    Вторая строка пары подрезов: нет профиля и длины, но задан угол второго запила.
    (Шаблон «Блок спецификации»: угол/сторона на следующей строке.)
    """
    if _cell_str(ws, r, cols["profile_code"]):
        return False
    if ws.cell(r, cols["length_mm"]).value is not None:
        return False
    try:
        _cell_int(ws, r, cols["cut_angle"])
    except Exception:
        return False
    return True


def parse_specification_with_stats(path: str | Path) -> tuple[list[SpecRow], ParseStats]:
    path = Path(path)
    logger.info("parse_specification start: %s", path)
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb.active
        hr = _find_header_row(ws)
        cols = _map_headers(ws, hr)

        item_col = cols.get("item_no")
        rows_out: list[SpecRow] = []
        last_item: int | None = None
        last_module = ""
        skipped_blank = 0
        skipped_invalid = 0

        r = hr + 1
        while r <= ws.max_row:
            profile = _cell_str(ws, r, cols["profile_code"])
            length_cell = ws.cell(r, cols["length_mm"]).value
            if profile is None and length_cell is None:
                skipped_blank += 1
                r += 1
                continue
            if not profile or length_cell is None:
                skipped_invalid += 1
                logger.warning(
                    "parse_specification skip row %d: incomplete profile/length (profile=%r, length=%r)",
                    r,
                    profile,
                    length_cell,
                )
                r += 1
                continue

            raw_item = _cell_str(ws, r, item_col) if item_col else None
            if raw_item:
                m = re.match(r"^(\d+)$", raw_item)
                if not m:
                    raise ValueError(f"Строка {r}: некорректный «№ п/п»: {raw_item!r}")
                last_item = int(m.group(1))
            elif last_item is None:
                raise ValueError(f"Строка {r}: отсутствует «№ п/п» в первой строке блока")

            mod = _cell_str(ws, r, cols["module_name"])
            if mod:
                last_module = mod
            if not last_module:
                raise ValueError(f"Строка {r}: не указано наименование модуля")

            try:
                length_mm = _cell_int(ws, r, cols["length_mm"])
                cut_angle = _cell_int(ws, r, cols["cut_angle"])
                qty = _cell_int(ws, r, cols["quantity"])
            except Exception as e:  # noqa: BLE001
                skipped_invalid += 1
                logger.warning("parse_specification skip row %d: bad numeric fields (%s)", r, e)
                r += 1
                continue
            if length_mm <= 0 or qty <= 0:
                skipped_invalid += 1
                logger.warning(
                    "parse_specification skip row %d: non-positive length/quantity (%d, %d)",
                    r,
                    length_mm,
                    qty,
                )
                r += 1
                continue

            qr = _cell_str(ws, r, cols.get("qr")) if "qr" in cols else None

            cut_angle_2: int | None = None
            row_advance = 1
            if r < ws.max_row and _is_second_angle_row(ws, r + 1, cols):
                try:
                    cut_angle_2 = _cell_int(ws, r + 1, cols["cut_angle"])
                    row_advance = 2
                except Exception as e:  # noqa: BLE001
                    logger.warning(
                        "parse_specification row %d: не удалось прочитать второй угол (%s)",
                        r + 1,
                        e,
                    )

            rows_out.append(
                SpecRow(
                    row_index=r,
                    item_no=last_item,
                    module_name=last_module.strip(),
                    profile_code=profile.strip(),
                    length_mm=length_mm,
                    cut_angle=cut_angle,
                    quantity=qty,
                    qr=qr,
                    cut_angle_2=cut_angle_2,
                )
            )
            r += row_advance

        if not rows_out:
            raise ValueError("Нет данных ниже заголовка")
        stats = ParseStats(
            parsed_rows=len(rows_out),
            skipped_blank_rows=skipped_blank,
            skipped_invalid_rows=skipped_invalid,
        )
        logger.info(
            "parse_specification done: rows=%d header_row=%d skipped_blank=%d skipped_invalid=%d",
            stats.parsed_rows,
            hr,
            stats.skipped_blank_rows,
            stats.skipped_invalid_rows,
        )
        return rows_out, stats
    finally:
        wb.close()


def parse_specification(path: str | Path) -> list[SpecRow]:
    rows, _stats = parse_specification_with_stats(path)
    return rows


def write_spec_workbook(path: str | Path, rows: list[tuple]) -> None:
    """Служебная запись тестовой спецификации. rows: список кортежей по колонкам."""
    path = Path(path)
    wb = Workbook()
    try:
        ws = wb.active
        assert ws is not None
        headers = list(HEADER_ALIASES.keys())
        for j, h in enumerate(headers, start=1):
            ws.cell(1, j, h)
        for i, row in enumerate(rows, start=2):
            for j, val in enumerate(row, start=1):
                ws.cell(i, j, val if val is not None else "")
        wb.save(path)
    finally:
        wb.close()
