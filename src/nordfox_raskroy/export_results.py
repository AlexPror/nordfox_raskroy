"""Экспорт результатов раскроя в Excel и PDF с цветами по модулям."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from nordfox_raskroy.models import CutEvent
from nordfox_raskroy.module_colors import (
    module_base_rgb,
    module_row_rgb,
    rgb_to_openpyxl_argb,
    rgb_to_pdf_hex,
)
from nordfox_raskroy.materials_library import row_mass_kg_display
from nordfox_raskroy.optimizer import format_cut_angles
from nordfox_raskroy.profile_codes import profile_label_for_code


def _opening_label(cut: CutEvent) -> str:
    return "склад" if cut.stock_opening_id == 0 else str(cut.stock_opening_id)


def _cut_rows(cuts: list[CutEvent]) -> list[tuple[str, ...]]:
    rows: list[tuple[str, ...]] = []
    for cut in cuts:
        d = cut.demand
        _, mtxt = row_mass_kg_display(d.profile_code, float(d.length_mm), 1.0)
        rows.append(
            (
                _opening_label(cut),
                d.module_name,
                d.profile_code,
                profile_label_for_code(d.profile_code),
                str(d.length_mm),
                format_cut_angles(d),
                "Обрезок" if cut.stock_source == "scrap" else "Новая",
                str(cut.stock_length_mm),
                str(cut.remainder_mm),
                mtxt,
            )
        )
    return rows


def _total_mass_footer(cuts: list[CutEvent]) -> tuple[str, bool]:
    """Текст итоговой массы и признак «есть хотя бы одна известная масса»."""
    total = 0.0
    any_known = False
    for cut in cuts:
        kg, _ = row_mass_kg_display(
            cut.demand.profile_code, float(cut.demand.length_mm), 1.0
        )
        if kg is not None:
            total += kg
            any_known = True
    if not any_known:
        return "—", False
    s = f"{total:.3f}".rstrip("0").rstrip(".")
    return s if s else "0", True


def export_cuts_excel(
    cuts: list[CutEvent],
    path: str | Path,
    *,
    summary: str = "",
) -> None:
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Раскрой"

    headers = (
        "Пруток №",
        "Модуль",
        "Тип профиля",
        "Серия",
        "Длина",
        "Угол",
        "Источник",
        "Заготовка мм",
        "Остаток мм",
        "Масса профиля, кг",
    )
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            patternType="solid",
            fgColor="FFE2E8F0",
        )

    for r, cut in enumerate(cuts, start=2):
        d = cut.demand
        rgb = module_row_rgb(d.module_name, is_scrap=cut.stock_source == "scrap")
        fill = PatternFill(
            patternType="solid",
            fgColor=rgb_to_openpyxl_argb(rgb),
        )
        _, mtxt = row_mass_kg_display(d.profile_code, float(d.length_mm), 1.0)
        vals = (
            _opening_label(cut),
            d.module_name,
            d.profile_code,
            profile_label_for_code(d.profile_code),
            str(d.length_mm),
            format_cut_angles(d),
            "Обрезок" if cut.stock_source == "scrap" else "Новая",
            str(cut.stock_length_mm),
            str(cut.remainder_mm),
            mtxt,
        )
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(r, c, val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    if cuts:
        tot_txt, tot_ok = _total_mass_footer(cuts)
        r_tot = len(cuts) + 2
        ws.merge_cells(start_row=r_tot, start_column=1, end_row=r_tot, end_column=9)
        c0 = ws.cell(r_tot, 1, "Итого масса профиля, кг")
        c0.font = Font(bold=True)
        c0.alignment = Alignment(horizontal="right", vertical="center")
        c0.fill = PatternFill(patternType="solid", fgColor="FFE2E8F0")
        ct = ws.cell(r_tot, 10, tot_txt if tot_ok else "—")
        ct.font = Font(bold=True)
        ct.alignment = Alignment(vertical="center")
        ct.fill = PatternFill(patternType="solid", fgColor="FFE2E8F0")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    leg = wb.create_sheet("Цвета модулей")
    leg.append(("Модуль", "Цвет (как в приложении)"))
    seen: set[str] = set()
    for cut in cuts:
        name = cut.demand.module_name
        if name in seen:
            continue
        seen.add(name)
        rgb = module_base_rgb(name)
        leg.append((name, rgb_to_openpyxl_argb(rgb)))
        row = leg.max_row
        assert row is not None
        for col in (1, 2):
            leg.cell(row, col).fill = PatternFill(
                patternType="solid",
                fgColor=rgb_to_openpyxl_argb(rgb),
            )

    if summary.strip():
        meta = wb.create_sheet("Сводка")
        meta.cell(1, 1, summary)
        meta.cell(1, 1).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)


def export_cuts_pdf(
    cuts: list[CutEvent],
    path: str | Path,
    *,
    summary: str = "",
    title: str = "Раскрой NordFox",
) -> None:
    path = Path(path)
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as e:  # pragma: no cover
        raise ImportError(
            "Для PDF установите: pip install reportlab"
        ) from e

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    story: list = [Paragraph(title, styles["Title"]), Spacer(1, 6 * mm)]
    if summary.strip():
        story.append(Paragraph(summary.replace("\n", "<br/>"), styles["Normal"]))
        story.append(Spacer(1, 6 * mm))

    headers = [
        "Пруток",
        "Модуль",
        "Профиль",
        "Серия",
        "Длина",
        "Угол",
        "Источник",
        "Загот. мм",
        "Остаток",
        "Масса профиля, кг",
    ]
    data: list[list[str]] = [headers]
    for cut in cuts:
        data.append(list(_cut_rows([cut])[0]))
    if cuts:
        tot_txt, tot_ok = _total_mass_footer(cuts)
        data.append(
            [
                "Итого масса профиля, кг",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                tot_txt if tot_ok else "—",
            ]
        )

    col_widths = [
        12 * mm,
        24 * mm,
        26 * mm,
        12 * mm,
        12 * mm,
        10 * mm,
        14 * mm,
        14 * mm,
        14 * mm,
        14 * mm,
    ]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    style_cmds: list[tuple] = [
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    for i, cut in enumerate(cuts, start=1):
        rgb = module_row_rgb(
            cut.demand.module_name,
            is_scrap=cut.stock_source == "scrap",
        )
        h = colors.HexColor(rgb_to_pdf_hex(rgb))
        style_cmds.append(("BACKGROUND", (0, i), (-1, i), h))
    if cuts:
        last = len(cuts) + 1
        style_cmds.append(("FONTNAME", (0, last), (-1, last), "Helvetica-Bold"))
        style_cmds.append(
            ("BACKGROUND", (0, last), (-1, last), colors.HexColor("#e2e8f0"))
        )

    t.setStyle(TableStyle(style_cmds))
    story.append(t)
    doc.build(story)
