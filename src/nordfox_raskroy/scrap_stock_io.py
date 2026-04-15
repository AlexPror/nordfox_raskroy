"""Импорт склада обрезков из Excel (длина × количество)."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook


def _norm(s: object | None) -> str:
    if s is None:
        return ""
    t = str(s).strip().lower()
    t = re.sub(r"[\s,]+", "", t)
    return t


def _find_scrap_header(ws, max_scan: int = 40) -> tuple[int, int, int] | None:
    """Строка и два индекса колонок: длина, количество."""
    for r in range(1, max_scan + 1):
        labels: dict[str, int] = {}
        for c in range(1, 24):
            key = _norm(ws.cell(r, c).value)
            if not key:
                continue
            labels[key] = c
        len_col = None
        qty_col = None
        for k, col in labels.items():
            if "длина" in k or k in ("l", "length"):
                len_col = col
            if "кол" in k or k.startswith("шт") or k in ("qty", "n", "колво"):
                qty_col = col
        if len_col and qty_col and len_col != qty_col:
            return (r, len_col, qty_col)
    return None


def _cell_int_row(ws, r: int, c: int) -> int | None:
    v = ws.cell(r, c).value
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        return int(round(v))
    s = str(v).strip().replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        return int(round(float(s)))
    except ValueError:
        return None


def parse_scrap_inventory(
    path: str | Path,
    *,
    max_pieces: int = 25_000,
) -> tuple[list[int], list[str]]:
    """
    Возвращает список длин каждого куска (повтор qty раз) и предупреждения.
    """
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb["Склад"] if "Склад" in wb.sheetnames else wb.active
        assert ws is not None
        return _read_scrap_sheet(ws, max_pieces)
    finally:
        close = getattr(wb, "close", None)
        if callable(close):
            close()


def _read_scrap_sheet(ws, max_pieces: int) -> tuple[list[int], list[str]]:
    found = _find_scrap_header(ws)
    if found is None:
        raise ValueError(
            "В файле склада не найдены колонки «Длина» и «Количество» "
            "(первые 40 строк активного листа)."
        )
    hr, c_len, c_qty = found
    pieces: list[int] = []
    warns: list[str] = []
    r = hr + 1
    while len(pieces) < max_pieces:
        ln = _cell_int_row(ws, r, c_len)
        qty = _cell_int_row(ws, r, c_qty)
        if ln is None and qty is None:
            break
        if ln is None or qty is None:
            warns.append(f"Строка {r}: пропуск (неполные длина/количество)")
            r += 1
            continue
        if ln <= 0 or qty <= 0:
            warns.append(f"Строка {r}: длина и количество должны быть > 0")
            r += 1
            continue
        for _ in range(qty):
            if len(pieces) >= max_pieces:
                warns.append(
                    f"Достигнут лимит {max_pieces} кусков; остальные строки не загружены"
                )
                return pieces, warns
            pieces.append(ln)
        r += 1

    if not pieces:
        raise ValueError("В файле склада нет ни одной строки с длиной и количеством")
    return pieces, warns
